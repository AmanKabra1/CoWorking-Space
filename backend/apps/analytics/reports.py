"""PDF and Excel report generators using ReportLab and openpyxl."""
import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

_DARK = colors.HexColor('#1a1a2e')
_BLUE = colors.HexColor('#4361ee')
_LIGHT = colors.HexColor('#f8f9fa')

_HEADER_STYLE = [
    ('BACKGROUND', (0, 0), (-1, 0), _DARK),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, _LIGHT]),
]


def _make_table(data, col_widths, align_right_cols=None):
    t = Table(data, colWidths=col_widths)
    style = list(_HEADER_STYLE)
    if align_right_cols:
        for col in align_right_cols:
            style.append(('ALIGN', (col, 0), (col, -1), 'RIGHT'))
    t.setStyle(TableStyle(style))
    return t


# ─── Revenue report ───────────────────────────────────────

def generate_revenue_pdf(data, start, end):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("CoWorkHub — Revenue Report", styles['Title']))
    elems.append(Paragraph(f"Period: {start} to {end}", styles['Normal']))
    elems.append(Spacer(1, 0.5 * cm))

    totals = data.get('totals', {})
    elems.append(_make_table(
        [
            ['Metric', 'Amount'],
            ['Total Invoiced', f"₹{float(totals.get('invoiced', 0)):,.2f}"],
            ['Total Paid', f"₹{float(totals.get('paid', 0)):,.2f}"],
            ['Invoice Count', str(totals.get('invoice_count', 0))],
        ],
        col_widths=[10 * cm, 6 * cm],
        align_right_cols=[1],
    ))
    elems.append(Spacer(1, 0.5 * cm))

    by_period = data.get('by_period', [])
    if by_period:
        elems.append(Paragraph("Revenue by Period", styles['Heading2']))
        rows = [['Period', 'Invoiced (₹)', 'Invoices']]
        for row in by_period:
            rows.append([row['period'], f"{float(row['invoiced']):,.2f}", str(row['invoice_count'])])
        elems.append(_make_table(rows, col_widths=[7 * cm, 6 * cm, 3 * cm], align_right_cols=[1, 2]))

    by_company = data.get('by_company', [])
    if by_company:
        elems.append(Spacer(1, 0.5 * cm))
        elems.append(Paragraph("Revenue by Company", styles['Heading2']))
        rows = [['Company', 'Invoiced (₹)', 'Invoices']]
        for row in by_company:
            rows.append([row['company'], f"{float(row['invoiced']):,.2f}", str(row['count'])])
        elems.append(_make_table(rows, col_widths=[8 * cm, 5 * cm, 3 * cm], align_right_cols=[1, 2]))

    doc.build(elems)
    buf.seek(0)
    return buf


def generate_revenue_excel(data, start, end):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    ws['A1'] = f"CoWorkHub — Revenue Report: {start} to {end}"
    ws['A1'].font = Font(bold=True, size=14)

    totals = data.get('totals', {})
    ws.append([])
    ws.append(['Summary', ''])
    ws.append(['Total Invoiced', float(totals.get('invoiced', 0))])
    ws.append(['Total Paid', float(totals.get('paid', 0))])
    ws.append(['Invoice Count', totals.get('invoice_count', 0)])

    ws.append([])
    ws.append(['Period', 'Invoiced', 'Invoice Count'])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='1a1a2e')
        cell.font = Font(bold=True, color='FFFFFF')

    for row in data.get('by_period', []):
        ws.append([row['period'], float(row['invoiced']), row['invoice_count']])

    if data.get('by_company'):
        ws.append([])
        ws.append(['Company', 'Invoiced', 'Invoice Count'])
        for row in data['by_company']:
            ws.append([row['company'], float(row['invoiced']), row['count']])

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Booking report ───────────────────────────────────────

def generate_booking_pdf(data, start, end):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("CoWorkHub — Booking Report", styles['Title']))
    elems.append(Paragraph(f"Period: {start} to {end}", styles['Normal']))
    elems.append(Spacer(1, 0.3 * cm))
    elems.append(Paragraph(f"Total Bookings: {data.get('total', 0)}", styles['Normal']))
    elems.append(Paragraph(
        f"Avg Duration: {float(data.get('avg_duration_hours', 0)):.1f} hours",
        styles['Normal'],
    ))
    elems.append(Spacer(1, 0.5 * cm))

    by_status = data.get('by_status', {})
    if by_status:
        elems.append(Paragraph("Bookings by Status", styles['Heading2']))
        rows = [['Status', 'Count']] + [[k.replace('_', ' ').title(), str(v)] for k, v in by_status.items()]
        elems.append(_make_table(rows, col_widths=[10 * cm, 4 * cm]))
        elems.append(Spacer(1, 0.5 * cm))

    by_facility = data.get('by_facility', [])
    if by_facility:
        elems.append(Paragraph("Top Facilities", styles['Heading2']))
        rows = [['Facility', 'Bookings', 'Revenue (₹)']]
        for row in by_facility:
            rows.append([row['facility'], str(row['bookings']), f"{float(row['revenue']):,.2f}"])
        elems.append(_make_table(rows, col_widths=[8 * cm, 4 * cm, 4 * cm], align_right_cols=[2]))
        elems.append(Spacer(1, 0.5 * cm))

    by_weekday = data.get('by_weekday', [])
    if by_weekday:
        elems.append(Paragraph("Bookings by Day of Week", styles['Heading2']))
        rows = [['Day', 'Bookings']] + [[row['day'], str(row['bookings'])] for row in by_weekday]
        elems.append(_make_table(rows, col_widths=[10 * cm, 4 * cm]))

    doc.build(elems)
    buf.seek(0)
    return buf


def generate_booking_excel(data, start, end):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    ws['A1'] = f"CoWorkHub — Booking Report: {start} to {end}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(['Total Bookings', data.get('total', 0)])
    ws.append(['Avg Duration (hrs)', float(data.get('avg_duration_hours', 0))])

    ws.append([])
    ws.append(['Status', 'Count'])
    for k, v in data.get('by_status', {}).items():
        ws.append([k.replace('_', ' ').title(), v])

    ws.append([])
    ws.append(['Facility', 'Bookings', 'Revenue'])
    for row in data.get('by_facility', []):
        ws.append([row['facility'], row['bookings'], float(row['revenue'])])

    ws.append([])
    ws.append(['Day of Week', 'Bookings'])
    for row in data.get('by_weekday', []):
        ws.append([row['day'], row['bookings']])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
