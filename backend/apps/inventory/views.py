from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, extend_schema_view

from apps.accounts.permissions import IsSuperAdminOrCompanyAdmin
from apps.core.exporters import build_export_response
from .models import InventoryItem, StockMovement
from .serializers import (
    InventoryItemSerializer,
    StockMovementSerializer,
    StockAdjustSerializer,
)


def _leased_building_ids(user):
    """Building IDs the user's company occupies (via desks or parking)."""
    from apps.workspace.models import Desk, ParkingSlot
    if not user.company_id:
        return []
    desk_ids = Desk.objects.filter(
        company_id=user.company_id
    ).values_list('room__floor__building_id', flat=True)
    park_ids = ParkingSlot.objects.filter(
        company_id=user.company_id
    ).values_list('building_id', flat=True)
    return list(set(desk_ids) | set(park_ids))


@extend_schema_view(
    list=extend_schema(tags=['Inventory']),
    retrieve=extend_schema(tags=['Inventory']),
    create=extend_schema(tags=['Inventory']),
    update=extend_schema(tags=['Inventory']),
    partial_update=extend_schema(tags=['Inventory']),
    destroy=extend_schema(tags=['Inventory']),
)
class InventoryItemViewSet(viewsets.ModelViewSet):
    """
    Building inventory — pantry, canteen, water, appliances, etc.

    - Super Admin: all buildings.
    - Company Admin: only buildings their company occupies.
    Restock / consume via custom actions, which log a StockMovement.
    """

    serializer_class = InventoryItemSerializer
    permission_classes = [IsSuperAdminOrCompanyAdmin]
    filterset_fields = ['building', 'category', 'is_active']
    search_fields = ['name', 'notes']

    def get_queryset(self):
        user = self.request.user
        qs = InventoryItem.objects.select_related('building')
        if user.is_super_admin:
            return qs
        return qs.filter(building_id__in=_leased_building_ids(user))

    @extend_schema(tags=['Inventory'], request=StockAdjustSerializer, responses={200: InventoryItemSerializer})
    @action(detail=True, methods=['post'], url_path='restock')
    def restock(self, request, pk=None):
        """Add stock (stock-in) and log the movement."""
        return self._adjust(request, InventoryItem, StockMovement.IN)

    @extend_schema(tags=['Inventory'], request=StockAdjustSerializer, responses={200: InventoryItemSerializer})
    @action(detail=True, methods=['post'], url_path='consume')
    def consume(self, request, pk=None):
        """Remove stock (stock-out) and log the movement."""
        return self._adjust(request, InventoryItem, StockMovement.OUT)

    def _adjust(self, request, _model, direction):
        item = self.get_object()
        serializer = StockAdjustSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        qty = serializer.validated_data['quantity']
        reason = serializer.validated_data.get('reason', '')

        if direction == StockMovement.OUT:
            if qty > item.quantity:
                return Response(
                    {'detail': f'Cannot consume {qty}; only {item.quantity} {item.unit} in stock.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            item.quantity -= qty
        else:
            item.quantity += qty

        item.save(update_fields=['quantity', 'updated_at'])
        StockMovement.objects.create(
            item=item, direction=direction, quantity=qty,
            reason=reason, performed_by=request.user,
        )
        return Response(InventoryItemSerializer(item).data)

    @extend_schema(tags=['Inventory'], responses={200: InventoryItemSerializer(many=True)})
    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock(self, request):
        """Items at or below their reorder level."""
        items = [i for i in self.get_queryset() if i.is_low_stock]
        return Response(InventoryItemSerializer(items, many=True).data)

    @extend_schema(tags=['Inventory'], responses={200: StockMovementSerializer(many=True)})
    @action(detail=True, methods=['get'], url_path='movements')
    def movements(self, request, pk=None):
        """Stock-movement history for one item."""
        item = self.get_object()
        qs = item.movements.select_related('performed_by').all()
        return Response(StockMovementSerializer(qs, many=True).data)

    @extend_schema(tags=['Inventory'])
    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Download the inventory list as Excel / Word / PDF (?format=excel|word|pdf)."""
        items = self.filter_queryset(self.get_queryset())
        # 'ID' is first so an exported file can be edited and re-uploaded (import
        # matches rows back to records by this id). Don't remove the ID column.
        headers = ['ID', 'Item', 'Category', 'Building', 'Quantity', 'Unit', 'Reorder Level', 'Unit Cost', 'Low Stock']
        rows = [
            [
                str(i.id), i.name, i.get_category_display(), i.building.name,
                i.quantity, i.unit, i.reorder_level, i.unit_cost,
                'Yes' if i.is_low_stock else 'No',
            ]
            for i in items
        ]
        return build_export_response(
            request.query_params.get('fmt'),
            'inventory', 'CoWorkHub — Inventory', headers, rows,
        )

    @extend_schema(tags=['Inventory'])
    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        Bulk-update inventory from an edited export file (.xlsx). Upload the file
        as multipart 'file'. Rows are matched to records by the ID column; only
        items the user can access are touched. Rows without a matching ID are skipped.
        Editable columns: Item, Category, Quantity, Unit, Reorder Level, Unit Cost.
        """
        from decimal import Decimal, InvalidOperation
        from openpyxl import load_workbook

        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Upload an .xlsx file as "file".'}, status=status.HTTP_400_BAD_REQUEST)
        if not upload.name.lower().endswith('.xlsx'):
            return Response({'detail': 'Only .xlsx files are supported.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read the file — is it a valid .xlsx?'},
                            status=status.HTTP_400_BAD_REQUEST)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return Response({'detail': 'The file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        header = [str(c).strip() if c is not None else '' for c in all_rows[0]]
        col = {name: i for i, name in enumerate(header)}
        if 'ID' not in col:
            return Response({'detail': 'Missing the ID column — upload a file exported from this page.'},
                            status=status.HTTP_400_BAD_REQUEST)

        cat_by_label = {label: value for value, label in InventoryItem.CATEGORY_CHOICES}
        editable = self.get_queryset()
        updated, skipped = 0, 0

        def cell(row, name):
            i = col.get(name)
            return row[i] if i is not None and i < len(row) else None

        def to_decimal(v):
            try:
                return Decimal(str(v))
            except (InvalidOperation, TypeError, ValueError):
                return None

        for row in all_rows[1:]:
            raw_id = cell(row, 'ID')
            if not raw_id:
                skipped += 1
                continue
            item = editable.filter(id=str(raw_id).strip()).first()
            if not item:
                skipped += 1
                continue

            name = cell(row, 'Item')
            if name:
                item.name = str(name).strip()
            unit = cell(row, 'Unit')
            if unit:
                item.unit = str(unit).strip()
            cat = cell(row, 'Category')
            if cat and str(cat).strip() in cat_by_label:
                item.category = cat_by_label[str(cat).strip()]
            for field, header_name in [('quantity', 'Quantity'), ('reorder_level', 'Reorder Level'), ('unit_cost', 'Unit Cost')]:
                val = to_decimal(cell(row, header_name))
                if val is not None:
                    setattr(item, field, val)
            item.save()
            updated += 1

        return Response({'updated': updated, 'skipped': skipped})
